import base64
import io
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import load_workbook
from pypdf import PdfReader
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

BASE_URL = "https://www.fire.ca.gov"
UPDATES_URL = f"{BASE_URL}/incidents/2025/1/7/palisades-fire/updates"
OUTPUT_EXCEL = "3. palisades_fire_compiled.xlsx"
OUTPUT_SHEET = "Palisades_Updates"
NEW_COLUMNS_LOG = "NEW_COLUMNS_README.md"
WAIT_TIMEOUT_SECONDS = 20
INITIAL_WAIT_SECONDS = 10
PAGE_DELAY_SECONDS = 3

REQUIRED_SCHEMA = [
    "Report_Sequence",
    "Report_Title",
    "Incident_Name_Header",
    "Report_Date",
    "Report_Time",
    "Agencies_Listed_Top (social media links)",
    "Public_Information_Line",
    "Media_Line",
    "Online_Fire_Information",
    "Incident_Name",
    "Start_Date_Time",
    "Incident_Status",
    "Location",
    "Type",
    "Cause",
    "Counties",
    "Administration_Unit",
    "Unified_Command_Agency_Size",
    "Containment",
    "Structures_Threatened",
    "Structures_Destroyed",
    "Structures_Damaged",
    "Civilian_Injuries",
    "Firefighter_Injuries",
    "Civilian_Fatalities",
    "Firefighter_Fatalities",
    "Situation_Summary",
    "Operational_Update",
    "Evacuation_Zone_Change_Notice",
    "Initial_Protective_Actions",
    "Additional_Resources",
    "Incident_Demographics_Title",
    "Damage_Assessment_Text",
    "Disaster_Assessment",
    "Federal_Assistance_Text",
    "Disaster_Resource_Center_Text",
    "DRC_Westside",
    "DRC_Eastside",
    "DRC_Ventura",
    "Governors_Office_Press_Releases",
    "Family_Assistance_Text",
    "Missing_Persons_Hotline",
    "Public_Health_Information",
    "Operational_Briefings",
    "Community_Meeting",
    "Community_Meeting_Livestream",
    "Incident_Photos_Videos",
    "State_Park_Closures",
    "Evacuation_Order_Zones",
    "Evacuation_Order_Open_To_Residents",
    "Curfew_Order",
    "Evacuation_Warning_Zones",
    "Evacuation_Warning_Open_To_Residents",
    "Evacuation_Orders_Details",
    "Evacuation_Shelters",
    "Road_Closure_Sources",
    "Animal_Evacuation_Missing_Pets_Line",
    "Small_Animal_Shelters",
    "Large_Animal_Shelters",
    "Assigned_Resources_Text",
    "Engines",
    "Water_Tenders",
    "Helicopters",
    "Dozers",
    "Hand_Crews",
    "Other",
    "Total_Personnel",
    "Cooperating_Agencies",
    "PDF_File_Name",
    "PDF_Page_Count",
]

LABEL_MAPPING = {
    "Date": "Report_Date",
    "Time": "Report_Time",
    "Name": "Incident_Name",
    "Start Date/Time": "Start_Date_Time",
    "Incident Status": "Incident_Status",
    "Location": "Location",
    "Type": "Type",
    "Cause": "Cause",
    "Counties": "Counties",
    "Administration Unit": "Administration_Unit",
    "Size": "Unified_Command_Agency_Size",
    "Containment": "Containment",
    "Structures Threatened": "Structures_Threatened",
    "Structures Destroyed": "Structures_Destroyed",
    "Structures Damaged": "Structures_Damaged",
    "Civilian Injuries": "Civilian_Injuries",
    "Firefighter Injuries": "Firefighter_Injuries",
    "Civilian Fatalities": "Civilian_Fatalities",
    "Firefighter Fatalities": "Firefighter_Fatalities",
    "Situation Summary": "Situation_Summary",
    "Update": "Operational_Update",
    "Evacuation Zone Change Notice": "Evacuation_Zone_Change_Notice",
    "Initial Protective Actions": "Initial_Protective_Actions",
    "Additional Resources": "Additional_Resources",
    "Incident Demographics": "Incident_Demographics_Title",
    "Damage Assessment": "Damage_Assessment_Text",
    "Disaster Assessment": "Disaster_Assessment",
    "Federal Assistance": "Federal_Assistance_Text",
    "Disaster Resource Center": "Disaster_Resource_Center_Text",
    "Westside Location": "DRC_Westside",
    "Eastside Location": "DRC_Eastside",
    "Ventura County Location": "DRC_Ventura",
    "Ventura Location": "DRC_Ventura",
    "Governor's Office Press Releases": "Governors_Office_Press_Releases",
    "Governors Office Press Releases": "Governors_Office_Press_Releases",
    "Family Assistance Center": "Family_Assistance_Text",
    "Missing Persons Hotline": "Missing_Persons_Hotline",
    "Public Health Information": "Public_Health_Information",
    "Press Conferences and Operational Briefings": "Operational_Briefings",
    "Community Meeting": "Community_Meeting",
    "Community Meeting Livestream": "Community_Meeting_Livestream",
    "Incident Photos and Videos": "Incident_Photos_Videos",
    "State Park Closures": "State_Park_Closures",
    "The Following Zones are under mandatory evacuation order": "Evacuation_Order_Zones",
    "OPEN TO RESIDENTS ONLY!": "Evacuation_Order_Open_To_Residents",
    "Per the Los Angeles County Sheriff's Department": "Curfew_Order",
    "Evacuation Warning Zones": "Evacuation_Warning_Zones",
    "Evacuation Warning Open To Residents": "Evacuation_Warning_Open_To_Residents",
    "Evacuation Orders": "Evacuation_Orders_Details",
    "Evacuation Orders Details": "Evacuation_Orders_Details",
    "Evacuation Warnings": "Evacuation_Warning_Zones",
    "Evacuation Shelters": "Evacuation_Shelters",
    "Road Closures": "Road_Closure_Sources",
    "Missing Pets Line": "Animal_Evacuation_Missing_Pets_Line",
    "Small Animals": "Small_Animal_Shelters",
    "Large Animals": "Large_Animal_Shelters",
    "Initial Protection Actions": "Initial_Protective_Actions",
    "Repopulation/Recovery Information": "Additional_Resources",
    "Repopulation & Recovery Information": "Additional_Resources",
    "News Releases/Media Advisories": "Governors_Office_Press_Releases",
    "Unsafe Drinking Water Notice": "Public_Health_Information",
    "California Wildfires Reunifications": "Family_Assistance_Text",
    "Shelter in Place": "Initial_Protective_Actions",
    "Assigned Resources": "Assigned_Resources_Text",
    "Engines": "Engines",
    "Water Tenders": "Water_Tenders",
    "Helicopters": "Helicopters",
    "Dozers": "Dozers",
    "Hand Crews": "Hand_Crews",
    "Other": "Other",
    "Total Personnel": "Total_Personnel",
    "Cooperating Agencies": "Cooperating_Agencies",
}


def normalize_label(value: str) -> str:
    cleaned = value.replace("\xa0", " ").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.rstrip(":")


NORMALIZED_MAPPING = {normalize_label(key).lower(): value for key, value in LABEL_MAPPING.items()}
PRESERVE_FULL_LINE_COLUMNS = {
    "Public_Information_Line",
    "Media_Line",
    "Online_Fire_Information",
    "Missing_Persons_Hotline",
    "Animal_Evacuation_Missing_Pets_Line",
}
IGNORED_UNMAPPED_SECTIONS = {
    "evacuation information",
    "animal evacuation shelters",
    "current situation",
    "evacuation zones",
}
INLINE_SECTION_MAPPING = {
    "additional resources": "Additional_Resources",
    "disaster resource center": "Disaster_Resource_Center_Text",
    "missing pets line": "Animal_Evacuation_Missing_Pets_Line",
    "missing persons hotline": "Missing_Persons_Hotline",
    "the following zones are under mandatory evacuation order": "Evacuation_Order_Zones",
    "open to residents only!": "Evacuation_Order_Open_To_Residents",
    "per the los angeles county sheriff's department": "Curfew_Order",
    "evacuation warnings": "Evacuation_Warning_Zones",
    "livestream": "Community_Meeting_Livestream",
    "small animals": "Small_Animal_Shelters",
    "large animals": "Large_Animal_Shelters",
    "westside location": "DRC_Westside",
    "eastside location": "DRC_Eastside",
    "ventura county location": "DRC_Ventura",
    "ventura location": "DRC_Ventura",
}


def make_blank_report() -> dict[str, str]:
    return {column: "" for column in REQUIRED_SCHEMA}


BLOCK_LEVEL_TAGS = {
    "address", "article", "aside", "blockquote", "caption", "dd", "div",
    "dl", "dt", "figcaption", "figure", "footer", "form", "h1", "h2", "h3",
    "h4", "h5", "h6", "header", "hr", "li", "main", "nav", "ol", "p", "pre",
    "section", "table", "tbody", "td", "tfoot", "th", "thead", "tr", "ul",
}
PARAGRAPH_LEVEL_TAGS = {
    "p", "div", "li", "blockquote", "section", "article", "tr", "dd", "dt",
    "h1", "h2", "h3", "h4", "h5", "h6", "pre", "figure", "address",
}


def extract_visible_text(node: Tag) -> str:
    parts: list[str] = []

    def append_break(double: bool) -> None:
        if not parts:
            return
        joined = "".join(parts)
        trailing = len(joined) - len(joined.rstrip("\n"))
        needed = 2 if double else 1
        if trailing >= needed:
            return
        parts.append("\n" * (needed - trailing))

    def walk(current: Tag | NavigableString) -> None:
        if isinstance(current, NavigableString):
            text = str(current)
            if text:
                parts.append(text)
            return
        if not isinstance(current, Tag):
            return
        name = (current.name or "").lower()
        if name == "br":
            append_break(double=False)
            return
        # Skip <thead> (column-header scaffolding like "Name Phone Address URL")
        if name == "thead":
            return
        is_block = name in BLOCK_LEVEL_TAGS
        is_paragraph = name in PARAGRAPH_LEVEL_TAGS
        if is_block:
            append_break(double=is_paragraph)
        for child in current.children:
            walk(child)
        if is_block:
            append_break(double=is_paragraph)

    walk(node)
    raw = "".join(parts).replace("\r\n", "\n").replace("\xa0", " ")
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in raw.split("\n")]
    collapsed: list[str] = []
    blank_run = 0
    for line in lines:
        if line == "":
            blank_run += 1
            if blank_run <= 1:
                collapsed.append(line)
        else:
            blank_run = 0
            collapsed.append(line)
    while collapsed and collapsed[0] == "":
        collapsed.pop(0)
    while collapsed and collapsed[-1] == "":
        collapsed.pop()
    return "\n".join(collapsed)


HEADING_TAGS = {"h1", "h2", "h3", "h4", "h5", "h6"}


def extract_nodes_text(nodes, skip_headings: bool = False) -> str:
    chunks: list[str] = []
    for node in nodes:
        if isinstance(node, NavigableString):
            text = str(node).strip()
            if text:
                chunks.append(text)
            continue
        if not isinstance(node, Tag):
            continue
        if skip_headings and node.name in HEADING_TAGS:
            continue
        block_text = extract_visible_text(node)
        if block_text.strip():
            chunks.append(block_text)
    return "\n\n".join(chunks).strip()


def iter_section_blocks(section: Tag) -> list[tuple[str, list[Tag | NavigableString]]]:
    blocks: list[tuple[str, list[Tag | NavigableString]]] = []
    current_title = ""
    current_nodes: list[Tag | NavigableString] = []

    def flush_current() -> None:
        if not current_title:
            return
        blocks.append((current_title, current_nodes.copy()))

    for child in section.children:
        if isinstance(child, Tag) and child.name in {"h3", "h4", "h5"}:
            flush_current()
            current_title = child.get_text(" ", strip=True)
            current_nodes = []
            continue
        if current_title:
            current_nodes.append(child)
    flush_current()
    return blocks


def fill_inline_sections(block_text: str, report: dict[str, str]) -> None:
    if not block_text.strip():
        return

    current_column = ""
    current_lines: list[str] = []

    def flush_current() -> None:
        if not current_column:
            return
        value = "\n".join(current_lines).strip()
        if value and not report.get(current_column, "").strip():
            report[current_column] = value

    for raw_line in block_text.splitlines():
        line = raw_line.rstrip()
        normalized = normalize_label(line).lower()
        match_column = INLINE_SECTION_MAPPING.get(normalized)
        matched_label = normalized
        if not match_column:
            for label, mapped_column in INLINE_SECTION_MAPPING.items():
                if normalized.startswith(f"{label}:"):
                    match_column = mapped_column
                    matched_label = label
                    break
                if normalized.startswith(f"{label} "):
                    match_column = mapped_column
                    matched_label = label
                    break
        if match_column:
            flush_current()
            current_column = match_column
            current_lines = []
            if current_column in PRESERVE_FULL_LINE_COLUMNS and line.strip():
                current_lines.append(line.strip())
                continue
            label_pos = normalized.find(matched_label)
            trailing = line
            if label_pos != -1:
                trailing = line[label_pos + len(matched_label) :].strip()
            if trailing.startswith(":"):
                trailing = trailing[1:].strip()
            if trailing:
                current_lines.append(trailing)
            continue
        if current_column:
            current_lines.append(line)
    flush_current()


def extract_between(text: str, start_patterns: list[str], end_patterns: list[str]) -> str:
    start_match = None
    for pattern in start_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match and (start_match is None or match.start() < start_match.start()):
            start_match = match
    if start_match is None:
        return ""

    trailing_text = text[start_match.end() :]
    end_match = None
    for pattern in end_patterns:
        match = re.search(pattern, trailing_text, flags=re.IGNORECASE)
        if match and (end_match is None or match.start() < end_match.start()):
            end_match = match

    if end_match is None:
        return trailing_text.strip()
    return trailing_text[: end_match.start()].strip()


def fill_fallback_fields(content: Tag, report: dict[str, str]) -> None:
    text = extract_visible_text(content)
    fallback_ranges = {
        "Operational_Update": (
            [r"\bUpdate\b\s*:"],
            [r"\bInitial Protective Actions\b", r"\bIncident Demographics\b", r"\bAdditional Resources\b"],
        ),
        "Additional_Resources": (
            [r"\bAdditional Resources\b"],
            [r"\bIncident Demographics\b", r"\bFamily Assistance Center\b"],
        ),
        "Community_Meeting": (
            [r"\bCommunity Meeting\b"],
            [r"\bLivestream\b\s*:", r"\bIncident Photos and Videos\b", r"\bEvacuation Zones\b"],
        ),
        "Assigned_Resources_Text": (
            [r"\bAssigned Resources\b"],
            [r"\bEngines\b"],
        ),
        "Evacuation_Shelters": (
            [r"\bEvacuation Shelters\b"],
            [r"\bRoad Closures\b"],
        ),
        "Road_Closure_Sources": (
            [r"\bRoad Closures\b"],
            [r"\bAnimal Evacuation Shelters\b", r"\bAssigned Resources\b"],
        ),
        "Small_Animal_Shelters": (
            [r"\bSmall Animals\b\s*:"],
            [r"\bLarge Animals\b\s*:"],
        ),
        "Large_Animal_Shelters": (
            [r"\bLarge Animals\b\s*:"],
            [r"\bAssigned Resources\b"],
        ),
        "Evacuation_Orders_Details": (
            [r"\bEvacuation Orders\b"],
            [r"\bEvacuation Shelters\b", r"\bRoad Closures\b"],
        ),
        "Evacuation_Order_Zones": (
            [r"\bThe Following Zones are under mandatory evacuation order\b\s*:"],
            [r"\bOPEN TO RESIDENTS ONLY!\b", r"\bPer the Los Angeles County Sheriff's Department\b", r"\bEvacuation Warnings\b"],
        ),
        "Evacuation_Order_Open_To_Residents": (
            [r"\bOPEN TO RESIDENTS ONLY!\b"],
            [r"\bPalisades Fire Repopulation\b", r"\bPer the Los Angeles County Sheriff's Department\b", r"\bTo identify your evacuation zone\b", r"\bEvacuation Warnings\b"],
        ),
        "Curfew_Order": (
            [r"\bPer the Los Angeles County Sheriff's Department\b\s*:"],
            [r"\bTo identify your evacuation zone\b", r"\bTo receive updates for your evacuation zones\b", r"\bEvacuation Warnings\b", r"\bEvacuation Shelters\b"],
        ),
        "Evacuation_Warning_Zones": (
            [r"\bEvacuation Warnings\b"],
            [r"\bPalisades Fire Repopulation\b", r"\bPer the Los Angeles County Sheriff's Department\b", r"\bEvacuation Shelters\b"],
        ),
        "Evacuation_Warning_Open_To_Residents": (
            # Only fire when the OPEN TO RESIDENTS ONLY! marker appears after
            # the Evacuation Warnings heading. The lookahead positions the cut
            # at the marker itself so it is preserved verbatim in the capture.
            [r"\bEvacuation Warnings\b[\s\S]*?(?=\bOPEN TO RESIDENTS ONLY!)"],
            [r"\bPer the Los Angeles County Sheriff's Department\b", r"\bEvacuation Shelters\b", r"\bPalisades Fire Repopulation\b"],
        ),
        "Disaster_Assessment": (
            [r"\bDisaster Assessment\b"],
            [r"\bEvacuation Orders\b", r"\bEvacuation Shelters\b"],
        ),
        "Damage_Assessment_Text": (
            [r"\bDamage inspection teams\b"],
            [r"\bEvacuation Orders\b", r"\bEvacuation Shelters\b"],
        ),
        "Disaster_Resource_Center_Text": (
            [r"\bDisaster Resource Center\b\s*:"],
            [r"\bWestside Location\b", r"\bFamily Assistance Center\b"],
        ),
        "DRC_Westside": (
            [r"\bWestside Location\b"],
            [r"\bEastside Location\b", r"\bVentura Location\b", r"\bFamily Assistance Center\b"],
        ),
        "DRC_Eastside": (
            [r"\bEastside Location\b"],
            [r"\bVentura County Location\b", r"\bVentura Location\b", r"\bFamily Assistance Center\b"],
        ),
        "DRC_Ventura": (
            [r"\bVentura County Location\b", r"\bVentura Location\b"],
            [r"\bFamily Assistance Center\b"],
        ),
        "Animal_Evacuation_Missing_Pets_Line": (
            [r"\bMissing Pets Line\b\s*:"],
            [r"\bSmall Animals\b\s*:", r"\bLarge Animals\b\s*:"],
        ),
        "Community_Meeting_Livestream": (
            [r"\bLivestream\b\s*:"],
            [r"\bIncident Photos and Videos\b", r"\bEvacuation Zones\b"],
        ),
    }

    for column, (start_patterns, end_patterns) in fallback_ranges.items():
        if report.get(column, "").strip():
            continue
        extracted = extract_between(text, start_patterns, end_patterns)
        if extracted:
            report[column] = extracted

    line_fallbacks = {
        "Missing_Persons_Hotline": r"Missing Persons Hotline\s*:\s*([^\n]+)",
        "Evacuation_Zone_Change_Notice": r"(Please be advised:\s*Evacuation Zones have changed[^\n]*)",
        "Operational_Update": r"(Update:\s*[^\n]+(?:\n[^\n]+){0,5})",
        "Community_Meeting": r"(Community Meeting\s*[^\n]+)",
        "Evacuation_Order_Zones": r"(The Following Zones are under mandatory evacuation order:\s*[^\n]+)",
        "Evacuation_Order_Open_To_Residents": r"(OPEN TO RESIDENTS ONLY!\s*[^\n]+)",
        "Curfew_Order": r"(Per the Los Angeles County Sheriff's Department:\s*[^\n]+(?:\n[^\n]+){0,2})",
        "Evacuation_Warning_Open_To_Residents": r"Evacuation Warnings[\s\S]{0,300}?(OPEN TO RESIDENTS ONLY![^\n]*)",
        "Community_Meeting_Livestream": r"(Livestream:\s*[^\n]+)",
    }
    for column, pattern in line_fallbacks.items():
        if report.get(column, "").strip():
            continue
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            report[column] = match.group(1).strip()


def normalize_field_boundaries(report: dict[str, str]) -> None:
    initial_actions = report.get("Initial_Protective_Actions", "")
    additional_resources = report.get("Additional_Resources", "")
    marker = "Additional Resources"
    if initial_actions and marker in initial_actions:
        before, after = initial_actions.split(marker, 1)
        cleaned_before = before.strip()
        cleaned_after = f"{marker}{after}".strip()
        if cleaned_before:
            report["Initial_Protective_Actions"] = cleaned_before
        if cleaned_after and not additional_resources.strip():
            report["Additional_Resources"] = cleaned_after


def clean_field_text(value: str, keep_blank_lines: bool = False) -> str:
    if not value:
        return ""
    normalized = value.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in normalized.split("\n")]
    if keep_blank_lines:
        cleaned_lines = lines
    else:
        cleaned_lines = [line for line in lines if line]
    cleaned = "\n".join(cleaned_lines).strip()
    return re.sub(r"[ \t]{2,}", " ", cleaned)


def normalize_extracted_fields(report: dict[str, str]) -> None:
    # Trim all captured values.
    for key, value in list(report.items()):
        if isinstance(value, str):
            report[key] = value.strip()

    # Clean line breaks/empty lines for multiline fields.
    multiline_fields = [
        "Administration_Unit",
        "Situation_Summary",
        "Damage_Assessment_Text",
        "Federal_Assistance_Text",
        "Disaster_Resource_Center_Text",
        "DRC_Westside",
        "DRC_Eastside",
        "DRC_Ventura",
        "Family_Assistance_Text",
        "Governors_Office_Press_Releases",
        "Public_Health_Information",
        "Operational_Briefings",
        "Community_Meeting",
        "Community_Meeting_Livestream",
        "Evacuation_Order_Zones",
        "Evacuation_Order_Open_To_Residents",
        "Curfew_Order",
        "Evacuation_Warning_Zones",
        "Evacuation_Warning_Open_To_Residents",
        "Road_Closure_Sources",
        "Animal_Evacuation_Missing_Pets_Line",
        "Small_Animal_Shelters",
        "Large_Animal_Shelters",
        "Assigned_Resources_Text",
        "Cooperating_Agencies",
        "Additional_Resources",
        "Disaster_Assessment",
        "Initial_Protective_Actions",
    ]
    for field in multiline_fields:
        report[field] = clean_field_text(report.get(field, ""))

    blank_line_fields = ["Evacuation_Shelters", "Evacuation_Orders_Details"]
    for field in blank_line_fields:
        report[field] = clean_field_text(report.get(field, ""), keep_blank_lines=True)

    # Ensure one-line scalar values.
    scalar_one_line = [
        "Start_Date_Time",
        "Public_Information_Line",
        "Media_Line",
        "Online_Fire_Information",
        "Missing_Persons_Hotline",
        "Animal_Evacuation_Missing_Pets_Line",
    ]
    for field in scalar_one_line:
        report[field] = re.sub(r"\s+", " ", report.get(field, "")).strip()

    # Slice each field to stop before any downstream marker that belongs to a
    # different column.
    trim_before_markers = {
        "Situation_Summary": [
            "Update:", "Operational Update", "Initial Protective Actions",
            "Additional Resources", "Evacuation Zone Change Notice",
        ],
        "Operational_Update": [
            "Please be advised:", "Evacuation Zone Change Notice",
            "Initial Protective Actions", "Additional Resources", "Incident Demographics",
        ],
        "Federal_Assistance_Text": ["Disaster Resource Center"],
        "Family_Assistance_Text": ["Missing Persons Hotline"],
        "Operational_Briefings": ["Community Meeting", "Livestream:"],
        "Curfew_Order": [
            "To identify your evacuation zone",
            "To receive updates for your evacuation zones",
            "Evacuation Warnings",
            "Evacuation Shelters",
        ],
        "Evacuation_Warning_Zones": ["OPEN TO RESIDENTS ONLY!"],
        "Evacuation_Order_Open_To_Residents": [
            "Palisades Fire Repopulation", "Resident Access:",
            "Per the Los Angeles County Sheriff", "To identify your evacuation zone",
        ],
    }
    for field, markers in trim_before_markers.items():
        value = report.get(field, "")
        if not value:
            continue
        for marker in markers:
            pos = value.find(marker)
            if pos != -1:
                value = value[:pos].rstrip()
        report[field] = clean_field_text(value)

    # Slice each field to start at (and include) the earliest known start marker
    # so the section label travels with the content verbatim. When no marker
    # is found the captured text is kept as-is.
    slice_from_markers = {
        "Evacuation_Orders_Details": (
            [
                "Resident Access:",
                "Palisades Fire Repopulation",
                "Per the Los Angeles County Sheriff",
                "To identify your evacuation zone",
            ],
            True,
        ),
    }
    for field, (markers, keep_blank_lines) in slice_from_markers.items():
        value = report.get(field, "")
        if not value:
            continue
        positions = [value.find(marker) for marker in markers]
        positions = [pos for pos in positions if pos != -1]
        if positions:
            value = value[min(positions):]
        report[field] = clean_field_text(value, keep_blank_lines=keep_blank_lines)


def section_title_to_column(title: str) -> str:
    token = normalize_label(title)
    token = re.sub(r"[^A-Za-z0-9]+", "_", token).strip("_")
    return token if token else "Unmapped_Section"


def parse_datetime(report: dict[str, str]) -> datetime | None:
    date_text = report.get("Report_Date", "").strip()
    time_text = report.get("Report_Time", "").strip()
    if not date_text:
        return None
    combined = f"{date_text} {time_text}".strip()
    formats = ("%m/%d/%Y %I:%M %p", "%m/%d/%y %I:%M %p", "%m/%d/%Y", "%m/%d/%y")
    for value_format in formats:
        try:
            return datetime.strptime(combined, value_format)
        except ValueError:
            continue
    return None


# Characters Chrome strips from document.title when generating the default
# "Save as PDF" filename. All are replaced with a single underscore.
PDF_FILENAME_INVALID_CHARS = re.compile(r'[\\/:*?"<>|]')


def title_to_pdf_filename(page_title: str) -> str:
    sanitized = PDF_FILENAME_INVALID_CHARS.sub("_", page_title).strip()
    return f"{sanitized}.pdf" if sanitized else ""


def capture_pdf_metadata(driver: webdriver.Chrome) -> tuple[str, str]:
    file_name = title_to_pdf_filename(driver.title or "")
    try:
        result = driver.execute_cdp_cmd(
            "Page.printToPDF",
            {"printBackground": False, "preferCSSPageSize": True},
        )
    except Exception:
        return file_name, ""

    data = result.get("data") if isinstance(result, dict) else None
    if not data:
        return file_name, ""

    try:
        reader = PdfReader(io.BytesIO(base64.b64decode(data)))
        return file_name, str(len(reader.pages))
    except Exception:
        return file_name, ""


def wait_for_element_with_refresh(driver: webdriver.Chrome, locator: tuple[str, str]) -> None:
    try:
        WebDriverWait(driver, INITIAL_WAIT_SECONDS).until(EC.presence_of_element_located(locator))
        return
    except TimeoutException:
        driver.refresh()

    WebDriverWait(driver, WAIT_TIMEOUT_SECONDS).until(EC.presence_of_element_located(locator))


def get_update_links(driver: webdriver.Chrome) -> list[str]:
    driver.get(UPDATES_URL)
    wait_for_element_with_refresh(driver, (By.CSS_SELECTOR, "div.detail-page"))
    soup = BeautifulSoup(driver.page_source, "html.parser")
    detail = soup.find("div", class_="detail-page")
    if detail is None:
        return []

    links: list[str] = []
    seen: set[str] = set()
    for anchor in detail.find_all("a", href=True):
        href = anchor["href"].strip()
        if not href:
            continue
        full_url = urljoin(BASE_URL, href)
        if full_url in seen:
            continue
        seen.add(full_url)
        links.append(full_url)
    return links


def extract_top_metadata(content: Tag, report: dict[str, str]) -> None:
    header = content.find("h1")
    if header:
        header_spans = header.find_all("span")
        if header_spans:
            report["Report_Title"] = header_spans[0].get_text(strip=False).strip()
        if len(header_spans) > 1:
            report["Incident_Name_Header"] = header_spans[1].get_text(strip=False).strip()

    rows = content.find_all("div", class_="row")
    if len(rows) < 2:
        return

    header_cols = rows[1].find_all("div", class_="col-sm")
    if not header_cols:
        return

    agencies = "\n".join(
        value for value in (anchor.get_text(" ", strip=True) for anchor in header_cols[0].find_all("a")) if value
    )
    report["Agencies_Listed_Top (social media links)"] = agencies

    if len(header_cols) < 2:
        return

    for info in header_cols[1].find_all("p"):
        line_raw = info.get_text(" ", strip=True)
        line_text = normalize_label(line_raw).lower()
        link = info.find("a")
        if "public information line" in line_text:
            report["Public_Information_Line"] = line_raw
        elif "media line" in line_text:
            report["Media_Line"] = line_raw
        elif "online fire information" in line_text:
            report["Online_Fire_Information"] = line_raw


def map_or_track_new_column(
    label: str,
    value: str,
    report: dict[str, str],
    new_column_records: list[dict[str, str]],
    source_url: str,
) -> None:
    normalized = normalize_label(label).lower()
    if normalized in IGNORED_UNMAPPED_SECTIONS:
        return
    column = NORMALIZED_MAPPING.get(normalized)
    if column:
        if value:
            report[column] = value
        return

    if not value.strip():
        return

    new_column = section_title_to_column(label)
    if new_column in report and report[new_column].strip():
        return
    report[new_column] = value
    new_column_records.append(
        {"source_url": source_url, "section_title": normalize_label(label), "column_name": new_column}
    )


def parse_report(
    driver: webdriver.Chrome, report_url: str, new_column_records: list[dict[str, str]]
) -> dict[str, str] | None:
    driver.get(report_url)
    wait_for_element_with_refresh(driver, (By.CSS_SELECTOR, "main.main-content"))
    soup = BeautifulSoup(driver.page_source, "html.parser")
    content = soup.find("main", class_="main-content")
    if content is None:
        return None

    # Remove hidden text once before extracting any fields.
    for hidden in content.select(".visually-hidden"):
        hidden.decompose()

    # Remove "Print status update report" link
    for candidate in content.find_all(["a", "button"]):
        candidate_text = candidate.get_text(" ", strip=True)
        if candidate_text == "Print status update report":
            candidate.decompose()

    report = make_blank_report()
    extract_top_metadata(content, report)

    for dl in content.find_all("dl"):
        columns = dl.find_all("dt")
        values = dl.find_all("dd")
        for column, value in zip(columns, values):
            value_text = extract_visible_text(value)
            if not value_text.strip():
                continue
            map_or_track_new_column(column.get_text(" ", strip=True), value_text, report, new_column_records, report_url)

    for section in content.find_all("div", class_="p-3"):
        blocks = iter_section_blocks(section)
        if not blocks:
            heading = section.find(["h3", "h4", "h5"])
            if heading is None:
                continue
            body = extract_nodes_text(section.children, skip_headings=True)
            if not body:
                continue
            map_or_track_new_column(heading.get_text(" ", strip=True), body, report, new_column_records, report_url)
            fill_inline_sections(body, report)
            continue

        for heading_text, heading_nodes in blocks:
            body = extract_nodes_text(heading_nodes)
            if not body:
                continue
            map_or_track_new_column(heading_text, body, report, new_column_records, report_url)
            fill_inline_sections(body, report)

    fill_fallback_fields(content, report)
    normalize_field_boundaries(report)
    normalize_extracted_fields(report)

    file_name, page_count = capture_pdf_metadata(driver)
    report["PDF_File_Name"] = file_name
    report["PDF_Page_Count"] = page_count

    return report


def sort_reports_chronologically(reports: list[dict[str, str]]) -> list[dict[str, str]]:
    return sorted(
        reports,
        key=lambda report: (
            parse_datetime(report) or datetime.max,
            report.get("Report_Title", ""),
        ),
    )


def build_dataframe(reports: list[dict[str, str]], discovered_columns: list[str]) -> pd.DataFrame:
    all_columns = REQUIRED_SCHEMA + [column for column in discovered_columns if column not in REQUIRED_SCHEMA]
    normalized_rows: list[dict[str, str]] = []
    for row in reports:
        normalized = {column: row.get(column, "") for column in all_columns}
        normalized_rows.append(normalized)
    return pd.DataFrame(normalized_rows, columns=all_columns)


def validate_dataframe(df: pd.DataFrame) -> None:
    if df.empty:
        raise RuntimeError("No reports were extracted; output file was not written.")
    missing_columns = [column for column in REQUIRED_SCHEMA if column not in df.columns]
    if missing_columns:
        raise RuntimeError(f"Required columns are missing from output schema: {missing_columns}")

    expected = [f"{index:03d}" for index in range(1, len(df) + 1)]
    actual = df["Report_Sequence"].astype(str).tolist()
    if actual != expected:
        raise RuntimeError("Report_Sequence is not monotonic from 001..N.")


def write_new_columns_log(records: list[dict[str, str]]) -> None:
    if not records:
        return

    unique_records: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for record in records:
        key = (record["source_url"], record["section_title"], record["column_name"])
        if key in seen:
            continue
        seen.add(key)
        unique_records.append(record)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# New Columns Added During Extraction",
        "",
        f"Run timestamp: {timestamp}",
        "",
        "| Source URL | Unmapped Section Title | Added Column Name |",
        "| --- | --- | --- |",
    ]
    for record in unique_records:
        lines.append(f"| {record['source_url']} | {record['section_title']} | {record['column_name']} |")
    lines.append("")

    with open(NEW_COLUMNS_LOG, "w", encoding="utf-8") as log_file:
        log_file.write("\n".join(lines))


def write_dataframe_preserve_formatting(df: pd.DataFrame, excel_path: str, sheet_name: str) -> None:
    if not Path(excel_path).exists():
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)
        return

    workbook = load_workbook(excel_path)
    if sheet_name not in workbook.sheetnames:
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    sheet = workbook[sheet_name]
    existing_headers = [cell.value for cell in sheet[1] if cell.value]
    ordered_columns = existing_headers.copy()
    for column in df.columns:
        if column not in ordered_columns:
            ordered_columns.append(column)
            sheet.cell(row=1, column=len(ordered_columns), value=column)

    aligned_df = df.reindex(columns=ordered_columns)
    max_col = len(ordered_columns)

    # Clear previous values without touching formatting/styles.
    for row_index in range(2, sheet.max_row + 1):
        for col_index in range(1, max_col + 1):
            sheet.cell(row=row_index, column=col_index).value = None

    for row_offset, row_values in enumerate(aligned_df.itertuples(index=False, name=None), start=2):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_offset, column=col_index, value=value)

    workbook.save(excel_path)


def main() -> None:
    driver = webdriver.Chrome(options=Options())
    try:
        update_links = get_update_links(driver)
        if not update_links:
            raise RuntimeError("Unable to get update links from the page.")

        reports: list[dict[str, str]] = []
        new_column_records: list[dict[str, str]] = []
        discovered_columns: list[str] = []
        discovered_set: set[str] = set()

        for index, link in enumerate(update_links):
            if index > 0:
                time.sleep(PAGE_DELAY_SECONDS)
            report = parse_report(driver, link, new_column_records)
            if report is None:
                continue
            for column in report:
                if column not in REQUIRED_SCHEMA and column not in discovered_set:
                    discovered_set.add(column)
                    discovered_columns.append(column)
            reports.append(report)

        reports = sort_reports_chronologically(reports)
        for index, report in enumerate(reports, start=1):
            report["Report_Sequence"] = f"{index:03d}"

        df = build_dataframe(reports, discovered_columns)
        validate_dataframe(df)
        write_dataframe_preserve_formatting(df, OUTPUT_EXCEL, OUTPUT_SHEET)
        write_new_columns_log(new_column_records)
        print(f"Excel file written: {OUTPUT_EXCEL}")
        if new_column_records:
            print(f"New column log written: {NEW_COLUMNS_LOG}")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()